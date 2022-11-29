# python3
import mysql.connector
import openpyxl

mydb = mysql.connector.connect(
  host='localhost',
  user='root',
  password='dbrmdk56',
  database='trailer'
)

mycursor = mydb.cursor(prepared=True)

filename = 'userdata_221015.xlsx'


wb = openpyxl.load_workbook(filename)
ws = wb.active

person_list = []
person_id_list = []
date_list = []

sql2 = "INSERT INTO user (person_id, nickname, strength) VALUES (%s, %s, %s);"

sql3 = "INSERT INTO walkscore (person_id, mon, tue, wed, thu, fri, sat, sun) VALUES (%s, %s, %s, %s, %s, %s, %s, %s);"

sql = "INSERT INTO walkcount (person_id, walk_date, steps) VALUES (%s, %s, %s);"

walkcount = []

walkscore = []

for row in ws.iter_rows(min_row=1, min_col=2, max_col=15, max_row=1, values_only=True):
  for val in row:
    person_id_list.append(val)

    person_list.append((val, val, "중"))
    walkscore.append((val, 3,3,3,3,3,3,3))

for col in ws.iter_rows(min_row=2, min_col=1, max_col=1, max_row=22, values_only=True):
  for val in col:
    date_list.append(val)

for col_num, col in enumerate(ws.iter_cols(min_row=2, max_row=22, min_col=2, max_col=15, values_only=True)):
  for date_num, val in enumerate(date_list):
    walkcount.append((person_id_list[col_num], str(date_list[date_num])[:10], col[date_num]))
    

# for data in person_list:
#   print(data)
# # for data in walkcount:
# #   print(data)
mycursor.executemany(sql3, walkscore)
# # # for person in walkcount:
# # #   mycursor.executemany(sql, person)
mydb.commit()  

print(mycursor.rowcount, "record inserted.")